"""
Excel report generator using openpyxl.
Generates downloadable .xlsx reports for all 6 IFRS 9 report types.
"""
import io
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ecl import ECLResult
from app.models.staging import StagingResult
from app.models.macro import MacroScenario
from app.models.provision import ProvisionRun, ProvisionMovement, GLEntry
from app.models.segment import Segment


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2D6A9F", end_color="2D6A9F", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="EEF4FB", end_color="EEF4FB", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
META_FONT = Font(name="Calibri", italic=True, size=9, color="666666")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_title_block(ws, title: str, month: str) -> int:
    """Write title + metadata, return next row number."""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"FinSight ECL — {title}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Reporting Month: {month[:4]}-{month[4:]}    Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}    Currency: BDT Crore"
    ws["A2"].font = META_FONT
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 16

    return 4  # start data from row 4


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18


def _write_data_row(
    ws, row: int, values: list, alt: bool = False, total: bool = False
) -> None:
    fill = TOTAL_FILL if total else (ALT_ROW_FILL if alt else None)
    font = TOTAL_FONT if total else DATA_FONT
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if fill:
            cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        if isinstance(val, (int, float, Decimal)) and col > 1:
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")


def _fmt_decimal(v: Optional[Decimal], decimals: int = 2) -> float:
    if v is None:
        return 0.0
    return round(float(v), decimals)


def _fmt_pct(v: Optional[Decimal]) -> str:
    if v is None:
        return "0.00%"
    return f"{float(v):.2f}%"


# ---------------------------------------------------------------------------
# Individual report builders
# ---------------------------------------------------------------------------

async def _build_ecl_summary(wb: Workbook, month: str, db: AsyncSession) -> None:
    ws = wb.active
    ws.title = "ECL Summary"
    row = _write_title_block(ws, "ECL Portfolio Summary", month)

    # Query ECL results grouped by stage
    result2 = await db.execute(
        select(
            ECLResult.stage,
            func.count(ECLResult.ecl_id),
            func.sum(ECLResult.ead),
            func.sum(ECLResult.ecl_base),
            func.sum(ECLResult.ecl_optimistic),
            func.sum(ECLResult.ecl_pessimistic),
            func.sum(ECLResult.ecl_weighted),
            func.avg(ECLResult.pd_12m),
            func.avg(ECLResult.lgd),
        )
        .where(ECLResult.reporting_month == month)
        .group_by(ECLResult.stage)
        .order_by(ECLResult.stage)
    )

    headers = [
        "Stage", "Loan Count", "EAD (Cr)", "ECL Base (Cr)",
        "ECL Optimistic (Cr)", "ECL Pessimistic (Cr)",
        "ECL Weighted (Cr)", "Avg PD 12M (%)", "Avg LGD (%)",
    ]
    _write_header_row(ws, row, headers)
    row += 1

    total_loans = total_ead = total_ecl = 0
    stage_labels = {1: "Stage 1 (Performing)", 2: "Stage 2 (SICR)", 3: "Stage 3 (Default)"}

    for i, (stage, cnt, ead, ecl_b, ecl_o, ecl_p, ecl_w, avg_pd, avg_lgd) in enumerate(result2):
        _write_data_row(ws, row, [
            stage_labels.get(stage, f"Stage {stage}"),
            int(cnt or 0),
            _fmt_decimal(ead),
            _fmt_decimal(ecl_b),
            _fmt_decimal(ecl_o),
            _fmt_decimal(ecl_p),
            _fmt_decimal(ecl_w),
            _fmt_decimal(avg_pd, 4) * 100 if avg_pd else 0,
            _fmt_decimal(avg_lgd, 4) * 100 if avg_lgd else 0,
        ], alt=(i % 2 == 1))
        total_loans += int(cnt or 0)
        total_ead += float(ead or 0)
        total_ecl += float(ecl_w or 0)
        row += 1

    _write_data_row(ws, row, ["TOTAL", total_loans, round(total_ead, 2), "", "", "", round(total_ecl, 2), "", ""], total=True)
    _set_col_widths(ws, [28, 12, 14, 15, 18, 18, 16, 14, 12])


async def _build_staging_summary(wb: Workbook, month: str, db: AsyncSession) -> None:
    ws = wb.active
    ws.title = "Staging Summary"
    row = _write_title_block(ws, "Staging Classification Summary", month)

    result2 = await db.execute(
        select(
            StagingResult.stage,
            func.count(StagingResult.staging_id).label("cnt"),
        )
        .where(StagingResult.reporting_month == month)
        .group_by(StagingResult.stage)
        .order_by(StagingResult.stage)
    )
    stage_counts = {r.stage: r.cnt for r in result2}
    total_staged = sum(stage_counts.values())

    # Override count
    override_result = await db.execute(
        select(
            StagingResult.stage,
            func.count(StagingResult.staging_id).label("cnt"),
        )
        .where(
            StagingResult.reporting_month == month,
            StagingResult.override_flag == True,  # noqa: E712
        )
        .group_by(StagingResult.stage)
    )
    override_counts = {r.stage: r.cnt for r in override_result}

    headers = ["Stage", "Description", "Loan Count", "% of Total", "Override Count", "Override %"]
    _write_header_row(ws, row, headers)
    row += 1

    stage_descs = {1: "12-Month ECL (Performing)", 2: "Lifetime ECL (SICR)", 3: "Lifetime ECL (Credit-Impaired)"}
    for i, (stage, cnt) in enumerate(sorted(stage_counts.items())):
        pct = (cnt / total_staged * 100) if total_staged else 0
        ov = override_counts.get(stage, 0)
        ov_pct = (ov / cnt * 100) if cnt else 0
        _write_data_row(ws, row, [
            f"Stage {stage}",
            stage_descs.get(stage, ""),
            cnt,
            f"{pct:.1f}%",
            ov,
            f"{ov_pct:.1f}%",
        ], alt=(i % 2 == 1))
        row += 1

    _write_data_row(ws, row, ["TOTAL", "", total_staged, "100.0%", sum(override_counts.values()), ""], total=True)
    _set_col_widths(ws, [10, 35, 12, 12, 15, 12])


async def _build_macro_sensitivity(wb: Workbook, month: str, db: AsyncSession) -> None:
    ws = wb.active
    ws.title = "Macro Sensitivity"
    row = _write_title_block(ws, "Macro Scenario Sensitivity", month)

    scenarios_result = await db.execute(
        select(MacroScenario)
        .where(MacroScenario.reporting_month == month)
        .order_by(MacroScenario.scenario_name)
    )
    scenarios = scenarios_result.scalars().all()

    # ECL totals per scenario multiplier
    ecl_result = await db.execute(
        select(
            func.sum(ECLResult.ecl_base),
            func.sum(ECLResult.ecl_optimistic),
            func.sum(ECLResult.ecl_pessimistic),
            func.sum(ECLResult.ecl_weighted),
        )
        .where(ECLResult.reporting_month == month)
    )
    ecl_row = ecl_result.one()
    ecl_base_total = float(ecl_row[0] or 0)
    ecl_opt_total = float(ecl_row[1] or 0)
    ecl_pess_total = float(ecl_row[2] or 0)
    ecl_w_total = float(ecl_row[3] or 0)

    # Scenario table
    headers = [
        "Scenario", "Status", "Weight (%)", "GDP Growth (%)", "CPI Inflation (%)",
        "BDT/USD Rate", "BB Repo Rate (%)", "Macro Multiplier", "ECL Contribution (Cr)",
    ]
    _write_header_row(ws, row, headers)
    row += 1

    scenario_ecl = {"BASE": ecl_base_total, "OPTIMISTIC": ecl_opt_total, "PESSIMISTIC": ecl_pess_total}
    for i, s in enumerate(scenarios):
        ecl_contrib = scenario_ecl.get(s.scenario_name, 0) * float(s.weight)
        _write_data_row(ws, row, [
            s.scenario_name,
            s.status,
            f"{float(s.weight) * 100:.1f}%",
            f"{float(s.gdp_growth or 0) * 100:.2f}%" if s.gdp_growth else "—",
            f"{float(s.cpi_inflation or 0) * 100:.2f}%" if s.cpi_inflation else "—",
            f"{float(s.bdt_usd_rate or 0):.4f}" if s.bdt_usd_rate else "—",
            f"{float(s.bb_repo_rate or 0) * 100:.2f}%" if s.bb_repo_rate else "—",
            float(s.macro_multiplier),
            round(ecl_contrib, 2),
        ], alt=(i % 2 == 1))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Total Weighted ECL (Cr)").font = LABEL_FONT
    ws.cell(row=row, column=9, value=round(ecl_w_total, 2)).font = TOTAL_FONT
    _set_col_widths(ws, [14, 10, 12, 14, 16, 12, 16, 18, 22])


async def _build_bb_regulatory(wb: Workbook, month: str, db: AsyncSession) -> None:
    ws = wb.active
    ws.title = "BB Regulatory"
    row = _write_title_block(ws, "Bangladesh Bank Regulatory Provision Schedule", month)

    result = await db.execute(
        select(
            ECLResult.stage,
            func.sum(ECLResult.ead),
            func.sum(ECLResult.ecl_weighted),
        )
        .where(ECLResult.reporting_month == month)
        .group_by(ECLResult.stage)
        .order_by(ECLResult.stage)
    )

    # BB CL classification mapping
    cl_map = {1: "Standard (STD)", 2: "Special Mention Account (SMA) / Sub-Standard (SS)", 3: "Doubtful (DF) / Bad & Loss (BL)"}
    min_prov_pct = {1: 1.0, 2: 5.0, 3: 20.0}  # minimum provision % per CL class

    headers = [
        "IFRS 9 Stage", "BB CL Classification", "Outstanding (EAD) (Cr)",
        "IFRS 9 ECL Provision (Cr)", "Coverage Ratio (%)",
        "BB Minimum Provision %", "Regulatory Shortfall / (Surplus) (Cr)",
    ]
    _write_header_row(ws, row, headers)
    row += 1

    grand_ead = grand_ecl = 0.0
    for i, (stage, ead, ecl) in enumerate(result):
        ead_f = float(ead or 0)
        ecl_f = float(ecl or 0)
        cov = (ecl_f / ead_f * 100) if ead_f else 0
        min_pct = min_prov_pct.get(stage, 1.0)
        min_req = ead_f * min_pct / 100
        shortfall = min_req - ecl_f
        _write_data_row(ws, row, [
            f"Stage {stage}",
            cl_map.get(stage, ""),
            round(ead_f, 2),
            round(ecl_f, 2),
            f"{cov:.2f}%",
            f"{min_pct:.1f}%",
            round(shortfall, 2),
        ], alt=(i % 2 == 1))
        grand_ead += ead_f
        grand_ecl += ecl_f
        row += 1

    grand_cov = (grand_ecl / grand_ead * 100) if grand_ead else 0
    _write_data_row(ws, row, [
        "TOTAL", "", round(grand_ead, 2), round(grand_ecl, 2),
        f"{grand_cov:.2f}%", "", "",
    ], total=True)
    _set_col_widths(ws, [14, 45, 22, 24, 16, 22, 32])


async def _build_ifrs7_disclosure(wb: Workbook, month: str, db: AsyncSession) -> None:
    ws = wb.active
    ws.title = "IFRS 7 Disclosure"
    row = _write_title_block(ws, "IFRS 7 Financial Instruments — Credit Risk Disclosure", month)

    result = await db.execute(
        select(
            ECLResult.stage,
            func.count(ECLResult.ecl_id),
            func.sum(ECLResult.ead),
            func.sum(ECLResult.ecl_weighted),
            func.avg(ECLResult.pd_12m),
            func.avg(ECLResult.pd_lifetime),
            func.avg(ECLResult.lgd),
        )
        .where(ECLResult.reporting_month == month)
        .group_by(ECLResult.stage)
        .order_by(ECLResult.stage)
    )

    headers = [
        "Stage", "Description", "Number of Loans", "Gross EAD (Cr)",
        "ECL Allowance (Cr)", "ECL / EAD (%)",
        "Avg 12M PD (%)", "Avg Lifetime PD (%)", "Avg LGD (%)",
    ]
    _write_header_row(ws, row, headers)
    row += 1

    descriptions = {
        1: "12-month ECL — no significant increase in credit risk",
        2: "Lifetime ECL — significant increase in credit risk",
        3: "Lifetime ECL — credit-impaired assets",
    }
    total_cnt = total_ead = total_ecl = 0
    for i, (stage, cnt, ead, ecl, avg_pd12, avg_pdlt, avg_lgd) in enumerate(result):
        ead_f = float(ead or 0)
        ecl_f = float(ecl or 0)
        ecl_rate = (ecl_f / ead_f * 100) if ead_f else 0
        _write_data_row(ws, row, [
            f"Stage {stage}",
            descriptions.get(stage, ""),
            int(cnt or 0),
            round(ead_f, 2),
            round(ecl_f, 2),
            f"{ecl_rate:.2f}%",
            f"{float(avg_pd12 or 0) * 100:.4f}%",
            f"{float(avg_pdlt or 0) * 100:.4f}%",
            f"{float(avg_lgd or 0) * 100:.2f}%",
        ], alt=(i % 2 == 1))
        total_cnt += int(cnt or 0)
        total_ead += ead_f
        total_ecl += ecl_f
        row += 1

    total_ecl_rate = (total_ecl / total_ead * 100) if total_ead else 0
    _write_data_row(ws, row, [
        "TOTAL", "", total_cnt, round(total_ead, 2), round(total_ecl, 2),
        f"{total_ecl_rate:.2f}%", "", "", "",
    ], total=True)
    _set_col_widths(ws, [10, 50, 14, 16, 18, 14, 16, 18, 12])


async def _build_gl_summary(
    wb: Workbook, month: str, db: AsyncSession, run_id: Optional[str]
) -> None:
    ws = wb.active
    ws.title = "GL Entry Summary"
    row = _write_title_block(ws, "General Ledger Entry Summary", month)

    # Get the relevant run
    if run_id:
        run_filter = ProvisionRun.run_id == run_id
    else:
        # Use latest approved/locked run for the month
        latest_result = await db.execute(
            select(ProvisionRun)
            .where(
                ProvisionRun.reporting_month == month,
                ProvisionRun.status.in_(["LOCKED", "APPROVED", "DRAFT"]),
            )
            .order_by(ProvisionRun.initiated_at.desc())
            .limit(1)
        )
        latest = latest_result.scalar_one_or_none()
        run_filter = ProvisionRun.run_id == (latest.run_id if latest else "NONE")
        if latest:
            run_id = latest.run_id

    # Run metadata
    run_result = await db.execute(
        select(ProvisionRun).where(run_filter)
    )
    run = run_result.scalar_one_or_none()
    if run:
        ws.cell(row=3, column=1, value=f"Run ID: {run.run_id}    Status: {run.status}    Type: {run.run_type}").font = META_FONT
        ws.merge_cells("A3:G3")

    # GL entries
    gl_result = await db.execute(
        select(GLEntry)
        .where(GLEntry.run_id == run_id)
        .order_by(GLEntry.entry_date, GLEntry.entry_id)
    )
    entries = gl_result.scalars().all()

    headers = [
        "Entry ID", "Entry Date", "DR Account", "CR Account",
        "Amount (Cr)", "Currency", "Entry Type", "Description", "Posted",
    ]
    _write_header_row(ws, row, headers)
    row += 1

    for i, e in enumerate(entries):
        _write_data_row(ws, row, [
            e.entry_id[:8] + "…",
            str(e.entry_date),
            e.dr_account,
            e.cr_account,
            round(float(e.amount), 2),
            e.currency,
            e.entry_type,
            e.description or "",
            "Yes" if e.posted else "No",
        ], alt=(i % 2 == 1))
        row += 1

    if not entries:
        ws.cell(row=row, column=1, value="No GL entries found for this run.").font = META_FONT

    # Movement waterfall
    row += 2
    ws.cell(row=row, column=1, value="Provision Movement Waterfall").font = LABEL_FONT
    row += 1

    mv_result = await db.execute(
        select(ProvisionMovement).where(ProvisionMovement.run_id == run_id)
    )
    movements = mv_result.scalars().all()
    _write_header_row(ws, row, ["Movement Type", "Amount (Cr)", "Account Count", "Notes"])
    row += 1
    for i, m in enumerate(movements):
        _write_data_row(ws, row, [
            m.movement_type,
            round(float(m.amount), 2),
            m.account_count,
            m.notes or "",
        ], alt=(i % 2 == 1))
        row += 1

    _set_col_widths(ws, [14, 14, 22, 22, 14, 10, 22, 40, 8])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

REPORT_BUILDERS = {
    "ECL_SUMMARY": _build_ecl_summary,
    "STAGING_SUMMARY": _build_staging_summary,
    "MACRO_SENSITIVITY": _build_macro_sensitivity,
    "BB_REGULATORY": _build_bb_regulatory,
    "IFRS7_DISCLOSURE": _build_ifrs7_disclosure,
    "GL_SUMMARY": None,  # handled separately (needs run_id)
}


async def generate_excel_report(
    report_id: str,
    month: str,
    run_id: Optional[str],
    db: AsyncSession,
) -> bytes:
    """
    Build an openpyxl workbook for the given report type and return
    the raw bytes of the .xlsx file.
    """
    wb = Workbook()

    if report_id == "GL_SUMMARY":
        await _build_gl_summary(wb, month, db, run_id)
    else:
        builder = REPORT_BUILDERS.get(report_id)
        if builder is None:
            raise ValueError(f"Unknown report_id: {report_id}")
        await builder(wb, month, db)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
